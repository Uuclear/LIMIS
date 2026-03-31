from __future__ import annotations

from typing import Any, Sequence

from django.db.models import QuerySet
from django.http import HttpResponse


def export_to_excel(
    queryset: QuerySet,
    fields: Sequence[str],
    headers: Sequence[str],
    filename: str = 'export.xlsx',
) -> HttpResponse:
    try:
        return _excel_openpyxl(queryset, fields, headers, filename)
    except ImportError:
        return _csv_fallback(queryset, fields, headers, filename)


def export_to_pdf(
    template_name: str,
    context: dict[str, Any],
) -> HttpResponse:
    from django.template.loader import render_to_string
    from weasyprint import HTML

    html_string = render_to_string(template_name, context)
    pdf = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{context.get("filename", "report.pdf")}"'
    return response


def _excel_openpyxl(
    queryset: QuerySet,
    fields: Sequence[str],
    headers: Sequence[str],
    filename: str,
) -> HttpResponse:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, obj in enumerate(queryset.iterator(), 2):
        for col_idx, field in enumerate(fields, 1):
            value = _resolve_field(obj, field)
            ws.cell(row=row_idx, column=col_idx, value=value)

    buf = BytesIO()
    wb.save(buf)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_to_csv(
    queryset: QuerySet,
    fields: Sequence[str],
    headers: Sequence[str],
    filename: str = 'export.csv',
) -> HttpResponse:
    """
    导出为 UTF-8 CSV（Content-Type 含 utf-8-sig，便于 Excel 打开中文列）。
    """
    import csv

    if not filename.lower().endswith('.csv'):
        filename = f'{filename}.csv'
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(headers)

    for obj in queryset.iterator():
        writer.writerow([_resolve_field(obj, f) for f in fields])

    return response


def _csv_fallback(
    queryset: QuerySet,
    fields: Sequence[str],
    headers: Sequence[str],
    filename: str,
) -> HttpResponse:
    csv_filename = filename.rsplit('.', 1)[0] + '.csv'
    return export_to_csv(queryset, fields, headers, filename=csv_filename)


def _resolve_field(obj: Any, field: str) -> Any:
    value = obj
    for attr in field.split('__'):
        value = getattr(value, attr, '')
        if callable(value):
            value = value()
    return value if value is not None else ''
