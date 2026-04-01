from __future__ import annotations

import base64
import random
import string
from datetime import date, datetime
from typing import Any, BinaryIO

from django.db import transaction
from django.db.models import QuerySet
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from core.utils.barcode import generate_qrcode
from core.utils.numbering import NumberGenerator
from core.audit import log_business_event

from .models import Sample, SampleDisposal


VALID_TRANSITIONS: dict[str, set[str]] = {
    'pending': {'testing'},
    'testing': {'tested'},
    'tested': {'retained', 'returned', 'disposed'},
    'retained': {'disposed', 'returned'},
    'disposed': set(),
    'returned': set(),
}


def generate_sample_no(commission) -> str:
    return NumberGenerator.generate(
        prefix='YP',
        project_code=getattr(commission, 'commission_no', None),
    )


def generate_blind_no() -> str:
    chars = string.ascii_uppercase + string.digits
    random_part = ''.join(random.choices(chars, k=8))
    return f'BL-{random_part}'


@transaction.atomic
def create_samples_from_commission(commission_id: int) -> list[Sample]:
    from apps.commissions.models import Commission

    commission = Commission.objects.select_related('project').get(pk=commission_id)
    items = commission.items.all() if hasattr(commission, 'items') else []
    samples = []

    for item in items:
        for _ in range(getattr(item, 'quantity', 1)):
            sample = Sample.objects.create(
                sample_no=generate_sample_no(commission),
                blind_no=generate_blind_no(),
                commission=commission,
                # 样品名称来源：委托项目（CommissionItem）里的“检测对象/检测项目”字段。
                # 旧实现错误地读取了不存在的 `item.name`，导致 name 为空，进而触发必填校验问题。
                # 有些 demo/历史数据里 `test_object` 可能用 '—' 表示缺失，此时应回退到 `test_item`。
                name=(
                    (lambda v: v if str(v).strip() not in ('', '-', '—', '－') else '')(  # type: ignore[no-any-return]
                        getattr(item, 'test_object', ''),
                    )
                    or getattr(item, 'test_item', '')
                ),
                specification=getattr(item, 'specification', ''),
                grade=getattr(item, 'grade', ''),
                quantity=1,
                unit=getattr(item, 'unit', '个'),
                sampling_date=getattr(
                    commission, 'sampling_date', timezone.now().date(),
                ),
                received_date=getattr(
                    commission, 'received_date', timezone.now().date(),
                ),
                sampling_location=getattr(commission, 'sampling_location', ''),
            )
            samples.append(sample)

    return samples


SAMPLE_IMPORT_TEMPLATE_COLUMNS: tuple[str, ...] = (
    'name',
    'specification',
    'grade',
    'quantity',
    'unit',
    'sampling_date',
    'received_date',
    'production_date',
    'sampling_location',
    'remark',
)

# openpyxl 表头别名（列名不区分大小写，去首尾空格）
_SAMPLE_HEADER_ALIASES: dict[str, frozenset[str]] = {
    'name': frozenset({'name', '样品名称'}),
    'specification': frozenset({'specification', '规格型号'}),
    'grade': frozenset({'grade', '设计强度', '设计等级', '等级'}),
    'quantity': frozenset({'quantity', '数量'}),
    'unit': frozenset({'unit', '单位'}),
    'sampling_date': frozenset({'sampling_date', '取样日期'}),
    'received_date': frozenset({'received_date', '收样日期', '接收日期'}),
    'production_date': frozenset({'production_date', '生产成型日期', '生产日期'}),
    'sampling_location': frozenset({'sampling_location', '取样地点'}),
    'remark': frozenset({'remark', '备注'}),
}


def _normalize_excel_header(cell: Any) -> str:
    s = str(cell or '').strip()
    if not s:
        return ''
    for field, aliases in _SAMPLE_HEADER_ALIASES.items():
        if s in aliases:
            return field
        low = s.lower()
        if low in aliases:
            return field
    return s.strip().lower().replace(' ', '_')


def _cell_to_date(val: Any):
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        from django.utils.dateparse import parse_date
        p = parse_date(val.strip())
        return p
    return None


def _cell_to_int(val: Any, default: int = 1) -> int:
    if val is None or val == '':
        return default
    if isinstance(val, bool):
        return int(val)
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    try:
        return int(str(val).strip())
    except ValueError:
        return default


def build_import_template_workbook():
    """返回 openpyxl Workbook（仅表头，与批量登记字段一致）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = 'samples'
    header_font = Font(bold=True)
    for col_idx, key in enumerate(SAMPLE_IMPORT_TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    return wb


def parse_samples_import_excel(
    file_obj: BinaryIO,
) -> tuple[list[tuple[int, dict[str, Any]]], list[dict[str, Any]]]:
    """
    解析批量导入 xlsx。
    返回 (rows, errors)；rows 为 (excel_row_index_1based, row_dict)；
    errors 元素为 {row, message}，row 为 Excel 行号（1-based）或 0 表示整表。
    """
    from openpyxl import load_workbook

    rows_out: list[tuple[int, dict[str, Any]]] = []
    errors: list[dict[str, Any]] = []

    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — 用户文件可能损坏
        return [], [{'row': 0, 'message': f'无法读取 Excel 文件: {e}'}]

    try:
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return [], [{'row': 1, 'message': '空表或缺少表头'}]

        col_index: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            field = _normalize_excel_header(cell)
            if not field:
                continue
            if field not in SAMPLE_IMPORT_TEMPLATE_COLUMNS:
                continue
            if field not in col_index:
                col_index[field] = idx

        required_headers = {'name', 'sampling_date', 'received_date'}
        missing = required_headers - set(col_index.keys())
        if missing:
            return [], [{
                'row': 1,
                'message': f'表头缺少必填列: {", ".join(sorted(missing))}',
            }]

        for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None:
                continue
            if not any(v not in (None, '') for v in row):
                continue

            raw = {}
            for field, cidx in col_index.items():
                val = row[cidx] if cidx < len(row) else None
                raw[field] = val

            entry: dict[str, Any] = {
                'name': str(raw.get('name') or '').strip(),
                'specification': str(raw.get('specification') or '').strip(),
                'grade': str(raw.get('grade') or '').strip(),
                'quantity': _cell_to_int(raw.get('quantity'), 1),
                'unit': str(raw.get('unit') or '个').strip() or '个',
                'sampling_date': _cell_to_date(raw.get('sampling_date')),
                'received_date': _cell_to_date(raw.get('received_date')),
                'production_date': _cell_to_date(raw.get('production_date')),
                'sampling_location': str(raw.get('sampling_location') or '').strip(),
                'remark': str(raw.get('remark') or '').strip(),
            }
            rows_out.append((excel_row_idx, entry))
    finally:
        wb.close()

    if not rows_out:
        return [], [{'row': 0, 'message': '没有可导入的数据行（请从第2行填写）'}]

    return rows_out, errors


@transaction.atomic
def create_samples_from_rows(commission_id: int, rows: list[dict[str, Any]]) -> list[Sample]:
    from apps.commissions.models import Commission

    commission = Commission.objects.select_related('project').get(pk=commission_id)
    samples: list[Sample] = []
    for row in rows:
        prod = row.get('production_date')
        sample = Sample.objects.create(
            sample_no=generate_sample_no(commission),
            blind_no=generate_blind_no(),
            commission=commission,
            name=row['name'],
            specification=row.get('specification', ''),
            grade=row.get('grade', ''),
            quantity=row.get('quantity', 1),
            unit=row.get('unit', '个'),
            sampling_date=row['sampling_date'],
            received_date=row['received_date'],
            production_date=prod,
            sampling_location=row.get('sampling_location', ''),
            remark=row.get('remark', ''),
        )
        samples.append(sample)
    return samples


@transaction.atomic
def change_sample_status(
    sample_id: int,
    new_status: str,
    user,
) -> Sample:
    sample = Sample.objects.select_for_update().get(pk=sample_id)
    if new_status == sample.status:
        return sample
    allowed = VALID_TRANSITIONS.get(sample.status, set())

    if new_status not in allowed:
        current_display = sample.get_status_display()
        target_display = dict(Sample.STATUS_CHOICES).get(new_status, new_status)
        raise ValidationError(
            f'样品状态不能从「{current_display}」变更为「{target_display}」'
        )

    sample.status = new_status
    sample.save(update_fields=['status', 'updated_at'])

    _log_status_change(sample, new_status, user)
    log_business_event(
        user=user,
        module='sample',
        action='status_change',
        entity='sample',
        entity_id=sample.pk,
        path=f'/api/v1/samples/samples/{sample.pk}/change-status/',
        payload={
            'sample_no': sample.sample_no,
            'new_status': new_status,
        },
    )
    return sample


def _log_status_change(sample: Sample, new_status: str, user) -> None:
    try:
        from apps.system.models import AuditLog
        AuditLog.objects.create(
            user=user if user and user.is_authenticated else None,
            username=getattr(user, 'username', ''),
            method='STATUS_CHANGE',
            path=f'/samples/{sample.pk}/change_status/',
            body=f'{{"sample_no":"{sample.sample_no}","new_status":"{new_status}"}}',
            ip_address=None,
            status_code=200,
        )
    except Exception:
        pass


def get_sample_timeline(sample_id: int) -> list[dict[str, Any]]:
    try:
        from apps.system.models import AuditLog
    except ImportError:
        return []

    logs = AuditLog.objects.filter(
        path__contains=f'/samples/{sample_id}/',
    ).order_by('timestamp')

    return [
        {
            'timestamp': log.timestamp,
            'user': log.username,
            'method': log.method,
            'action': _parse_action(log),
        }
        for log in logs
    ]


def _parse_action(log) -> str:
    if log.method == 'STATUS_CHANGE':
        return f'状态变更: {log.body}'
    return f'{log.method} {log.path}'


def get_retention_samples() -> QuerySet:
    return Sample.objects.filter(
        status='retained',
    ).select_related(
        'commission', 'commission__project',
    ).order_by('retention_deadline')


def generate_sample_label(sample_id: int) -> dict[str, Any]:
    sample = Sample.objects.select_related(
        'commission', 'commission__project',
    ).get(pk=sample_id)

    # 移动端可解析 `LIMIS:SAMPLE:{sample_no}` 后打开前端公开页 `/verify/sample/<sample_no>` 核对进度。
    qr_data = f'LIMIS:SAMPLE:{sample.sample_no}'
    qr_bytes = generate_qrcode(qr_data, size=200)
    qr_base64 = base64.b64encode(qr_bytes).decode('ascii')

    return {
        'sample_no': sample.sample_no,
        'blind_no': sample.blind_no or '',
        'name': sample.name,
        'specification': sample.specification,
        'grade': sample.grade,
        'commission_no': getattr(sample.commission, 'commission_no', ''),
        'project_name': (
            sample.commission.project.name
            if sample.commission and sample.commission.project
            else ''
        ),
        'sampling_date': str(sample.sampling_date),
        'received_date': str(sample.received_date),
        'qr_code': f'data:image/png;base64,{qr_base64}',
    }


@transaction.atomic
def dispose_sample(
    sample_id: int,
    disposal_type: str,
    disposal_date,
    handler,
    remark: str = '',
) -> SampleDisposal:
    sample = Sample.objects.select_for_update().get(pk=sample_id)
    target = 'returned' if disposal_type == 'return' else 'disposed'

    if sample.status == target:
        existing = (
            SampleDisposal.objects.filter(
                sample=sample,
                disposal_type=disposal_type,
                disposal_date=disposal_date,
            )
            .order_by('-id')
            .first()
        )
        if existing:
            return existing

    allowed = VALID_TRANSITIONS.get(sample.status, set())

    if target not in allowed:
        raise ValidationError(
            f'当前状态「{sample.get_status_display()}」不允许处置操作'
        )

    disposal = SampleDisposal.objects.create(
        sample=sample,
        disposal_type=disposal_type,
        disposal_date=disposal_date,
        handler=handler,
        remark=remark,
    )

    sample.status = target
    sample.disposal_date = disposal_date
    sample.disposal_method = dict(SampleDisposal.DISPOSAL_TYPE_CHOICES).get(
        disposal_type, disposal_type,
    )
    sample.save(update_fields=[
        'status', 'disposal_date', 'disposal_method', 'updated_at',
    ])

    _log_status_change(sample, target, handler)
    log_business_event(
        user=handler,
        module='sample',
        action='dispose',
        entity='sample',
        entity_id=sample.pk,
        path=f'/api/v1/samples/samples/{sample.pk}/dispose/',
        payload={
            'sample_no': sample.sample_no,
            'disposal_type': disposal_type,
            'target_status': target,
        },
    )
    return disposal
